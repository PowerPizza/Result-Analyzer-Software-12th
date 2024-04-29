"""
Date - 02-08-2023
Developer - scihack/powerpizza
Purpose - to sort the txt unsorted data in a excel file.
"""
import tkinter, json
from tkinter import *
from tkinter import filedialog, messagebox
from string_helper_funcs import *
from other_functions import *
import openpyxl
from openpyxl.styles import Alignment, PatternFill
import webbrowser
from pandas import DataFrame, to_numeric, ExcelWriter
from pandas.io.formats import excel

print("Starting . . . .")
root = tkinter.Tk()
root.geometry("900x600")
root.title("Result Analyzer (For 12th)")
root.state("zoomed")
root.iconphoto(True, PhotoImage(file="software_icon.png", master=root))

# --------------- constants and variables ------------
selected_file = None
subject_by_code = json.load(open("subject_code_refer.json", "r"))
prv_next_step = 2
configs_ = json.load(open("configs.json", "r"))
file_lines = None
all_processed_data = None
status_area_frame = None
excel.ExcelFormatter.header_style = None
df_ = DataFrame([[0, 0, 0], [0, 0, 0]])
unique_sub_codes = configs_["dominant_subjects"]
logs_string = ""
show_loading = False
# --------------------- END -----------------------

# ----------------- functions of DRY --------------
def add_log(message):
    global logs_string
    logs_string += message + "\n"
    logs_opt.config(background="#FF0000", foreground="#FFFFFF")
def filterDataLine1(line):
    to_ret = {"roll_no": int(line[0]), "gender": None, "name": "", "subject_codes": [], "result": "",
              "comp_subjects": []}

    # extracting gender, name and subject code format
    name_done = False
    for xo in range(1, len(line)):
        if line[xo] == "COMP":
            break

        if line[xo] == 'M':
            to_ret["gender"] = "MALE"

        elif line[xo] == 'F':
            to_ret["gender"] = "FEMALE"

        elif line[xo].replace(" ", "").isnumeric():
            to_ret["subject_codes"].append(line[xo])
            name_done = True  # it's obvious by valid data format that if subject code is there so name should have been done.

        else:
            if not name_done:
                to_ret["name"] += " " + line[xo]
    to_ret["name"] = rm_extra_spaces(to_ret['name'])

    # extracting result pass or fail
    to_ret["result"] = line[-1]

    if "COMP" in line:
        to_ret["result"] = "COMP"
        comp_subjs = line[line.index("COMP") + 1:]
        for itm in comp_subjs:
            to_ret["comp_subjects"].append(f"{subject_by_code[itm]['Name']} ({itm})")

    return to_ret


def dataLineExtractor(data_file):
    """
    main motive of this function is to extract the datalines (DATA_LINE1 and DATA_LINE2) but it also
    calculates some values like (overall marks and their totals, best5 marks and their totals).
    """
    dataLinePairs = []
    data_line1_found = False
    data_line1 = []
    for line in data_file.readlines():
        line = list_formatter(str(rm_extra_spaces(line.replace("\n", ""))).split(" "))
        if len(line):
            if len(line[0]) in range(6, 11) and line[0].isnumeric():
                if len(data_line1):
                    add_log(f"⚠ Data line 2 not found for `{data_line1}` Due to which entry will be ignored.")
                    # messagebox.showwarning("Missing DATA_LINE 2", f"Data line 2 not found for {data_line1}\nDue to which entry will be ignored.")
                    # print("Data line 2 not found for : ", data_line1)
                data_line1 = line
                data_line1_found = True
            elif (len(line[0]) == 3 and line[0].isnumeric()) or ("ABST" in data_line1):
                if data_line1_found:
                    filtered_line1 = filterDataLine1(data_line1)
                    to_append_ = {"LINE1": data_line1, "LINE2": line, "SubCode_MG": {},
                                  "overall_max_marks": configs_["max_marks_1_subject"]*len(filtered_line1["subject_codes"]),
                                  "best5_max_marks": configs_["max_best_sub_range"]*configs_["max_marks_1_subject"],
                                  "best5_total": 0, "overall_total": 0}

                    idxr = 0
                    marks_list = []
                    for code_ in filtered_line1["subject_codes"]:
                        to_append_["SubCode_MG"][code_] = [line[idxr], line[idxr+1]]
                        if line[idxr].isnumeric():
                            to_append_["overall_total"] += int(line[idxr])
                            if code_ != configs_["main_subject_code"]:
                                marks_list.append(int(line[idxr]))
                        idxr += 2
                    if len(marks_list):
                        to_append_["best5_total"] = sum(sorted(marks_list, reverse=True)[:configs_["max_best_sub_range"]-1]) + int(to_append_["SubCode_MG"][configs_["main_subject_code"]][0])

                    dataLinePairs.append(to_append_)
                    data_line1_found = False
                    data_line1 = []
                else:
                    # messagebox.showwarning("Missing DATA_LINE 1", f"Data line 1 not found for {line}\nDue to which entry will be ignored.")
                    add_log(f"⚠ Data line 1 not found for `{line}` Due to which entry will be ignored.")

                    # print("data line 1 not found for ", line)
                # messagebox.showwarning("Invalid data line", f"Invalid dataline found it may have missing its pair.\n{line}")

    # print(dataLinePairs)
    initDataFrame(dataLinePairs)
    # print(df_)
    begin_status(data_file.name)
    # DataFrame Creation

def initDataFrame(dataLinePairs):
    global df_, unique_sub_codes
    for pair in dataLinePairs:
        filtered_line = filterDataLine1(pair["LINE1"])
        for sub_code in filtered_line["subject_codes"]:
            if sub_code not in unique_sub_codes:
                unique_sub_codes.append(sub_code)
    columns_ = []
    list(map(lambda sub_code: columns_.extend([f"MARKS", f"GRADES"]), unique_sub_codes))
    df_ = DataFrame(
        columns=["roll_no", "student_name", "gender"] + columns_ + ["Max Marks OA", "Total OA", "Percentage OA", "Average OA"] + ["Max Marks B5", "Total B5", "Percentage B5",
                                                                                                         "Average B5"] + [
                    "result", "compartment"])

    for itm in dataLinePairs:
        filtered_r1 = filterDataLine1(itm["LINE1"])
        MG_in_pattern = []
        for code_ in unique_sub_codes:
            if code_ in itm["SubCode_MG"]:
                MG_in_pattern.extend(itm["SubCode_MG"][code_])
            else:
                MG_in_pattern.extend(["", ""])

        percentage_overall = "%.2f" % (itm["overall_total"] / itm["overall_max_marks"] * 100)
        percentage_best5 = "%.2f" % (itm["best5_total"] / itm["best5_max_marks"] * 100)
        average_overall = "%0.2f" % (itm["overall_total"] / len(itm["SubCode_MG"]))
        average_best5 = "%0.2f" % (itm["best5_total"] / configs_["max_best_sub_range"])
        df_.loc[df_.index.size] = [filtered_r1["roll_no"], filtered_r1["name"],
                                   filtered_r1["gender"]] + MG_in_pattern + [itm["overall_max_marks"],
                                                                             itm["overall_total"], percentage_overall,
                                                                             average_overall,
                                                                             itm["best5_max_marks"], itm["best5_total"],
                                                                             percentage_best5, average_best5,
                                                                             filtered_r1["result"],
                                                                             ", ".join(filtered_r1["comp_subjects"])]

def writeToExcel():
    export_file = filedialog.asksaveasfile(filetypes=(("Excel file", "*.xlsx"), ("All files", "*.*")), defaultextension=".xlsx")
    if not export_file:
        return

    columns_ = []
    list(map(lambda sub_code: columns_.extend([f"MARKS", f"GRADES"]), unique_sub_codes))
    merged_column_excel_format = ["", "", ""] + columns_ + (["Max Marks", "Total", "Percentage", "Average"]*2) + ["", ""]

    # -------------- CREATING OVER ALL SHEET ----------
    df_over_all = DataFrame(df_)
    df_over_all["MARKS"] = df_over_all["MARKS"].replace(['', ' '], None)
    df_over_all["MARKS"] = df_over_all["MARKS"].apply(lambda row_: to_numeric(row_, errors='coerce'))  # changed datatype of all MARKS columns to integer
    sub_total = df_over_all["MARKS"].sum()
    all_avg = df_over_all["MARKS"].sum() / (len(df_over_all["MARKS"].index) - df_over_all["MARKS"].isna().sum())
    df_over_all.loc["SUB_TOTAL", "MARKS"] = sub_total
    df_over_all.loc["ALL_AVERAGE", "MARKS"] = all_avg.__round__(2)
    excel_writer = ExcelWriter(export_file.name)
    df_over_all.columns = merged_column_excel_format
    df_over_all.to_excel(excel_writer, index=False, sheet_name="OverAll Result")

    # -------------- EXTRACTING TOP 10 STUDENTS IN A NEW SHEET -------
    df_top_10 = DataFrame(df_)
    df_top_10 = df_top_10.sort_values("Total B5", ascending=False).iloc[0:10]
    df_top_10.columns = merged_column_excel_format
    df_top_10.to_excel(excel_writer, index=False, sheet_name="Top 10")

    # ------------- CREATING SHEET OF EACH SINGLE SUBJECT ------------
    df_over_all = DataFrame(df_)
    df_over_all.columns = list(range(0, len(df_over_all.columns)))
    df_col_idxr = 3
    for uni_subcode in unique_sub_codes:
        df_single = DataFrame(df_over_all[[0, 1, 2, df_col_idxr, df_col_idxr+1]])
        df_single.loc[:, df_col_idxr] = to_numeric(df_single[df_col_idxr], errors='coerce')
        avg_marks = df_single[df_col_idxr].sum()
        if avg_marks != 0:
            avg_marks = "%0.2f" % (avg_marks / (len(df_single.index)-df_single[df_col_idxr].isna().sum()))  # .isna().sum() counts NaN values
        df_single.columns = ["Roll No.", "Name", "Gender", "Marks Obtained", "Grade"]
        df_single['Grade'] = df_single['Grade'].replace(['', ' '], None)
        df_single = df_single.dropna(subset=['Marks Obtained', 'Grade'], how='all')
        df_single = df_single.sort_values("Marks Obtained", ascending=False)
        df_single.insert(0, "S. no.", "")
        df_single['S. no.'] = range(1, len(df_single.index)+1)
        df_single.loc["average", ['Marks Obtained']] = f"AVG : {avg_marks}"
        df_single.to_excel(excel_writer, sheet_name=subject_by_code[uni_subcode]["Name"], index=False)
        df_col_idxr += 2

    # --------- SEPARATING TOPPERS (1ST 2ND 3RD.....nTH) IN NEW EXCEL SHEETS ------------
    df_cpy = DataFrame(df_)
    df_cpy["Percentage B5"] = to_numeric(df_cpy["Percentage B5"], errors="coerce")
    df_cpy = df_cpy.sort_values("Percentage B5", ascending=False)
    b5_perc_col_idx = list(df_cpy.columns).index("Percentage B5")

    df_cpy.columns = merged_column_excel_format
    sheet_names = {"95 PLUS": {"r1": 95, "r2": 100}, "Between 90 to 95": {"r1": 90, "r2": 95}, "Between 85 to 90": {"r1": 85, "r2": 90}, "Between 70 to 85": {"r1": 75, "r2": 85}, "Between 65 to 70": {"r1": 65, "r2": 70}, "Between 45 to 65": {"r1": 45, "r2": 65}, "Below 40": {"r1": 0, "r2": 40}}
    for limit_ in sheet_names:
        range_sheet = df_cpy[(df_cpy.iloc[:, b5_perc_col_idx] >= sheet_names[limit_]["r1"]) & (df_cpy.iloc[:, b5_perc_col_idx] < sheet_names[limit_]["r2"])]
        range_sheet.to_excel(excel_writer, sheet_name=limit_, index=False)

    # --------- ADDING HEADER WITH MERGED CELLS IN EXCEL FILE -----------
    excel_writer.close()  # necessary!! since I am not using with block.
    wb = openpyxl.load_workbook(export_file.name)

    def createOverAllSheetLayout(sheet_name):
        ws_ = wb[sheet_name]

        ws_.insert_rows(1, amount=1)
        row_pointer = 1

        table_header = [
            ["Roll no", 0],
            ["Student Name", 0],
            ["Gender", 0],
        ]
        table_header += map(lambda sub_name: [sub_name, 2], subcode_to_subname(list(unique_sub_codes)))
        table_header += [
            ["Over All", 4],
            [f"Best(5)(Best 4 + {subject_by_code[configs_['main_subject_code']]['Name']})", 4],
            ["Result", 0],
            ["Compartment", 0]
        ]

        merge_step = 0
        for i in range(len(table_header)):
            cell_ = ws_.cell(row=row_pointer, column=i + merge_step + 1, value=table_header[i][0])
            cell_.alignment = Alignment(horizontal="center")
            if table_header[i][1]:
                ws_.merge_cells(start_row=row_pointer, end_row=row_pointer, start_column=cell_.col_idx,
                               end_column=cell_.col_idx + table_header[i][1] - 1)
                merge_step += table_header[i][1] - 1
    createOverAllSheetLayout("OverAll Result")
    createOverAllSheetLayout("Top 10")
    for btw_sheet in sheet_names:
        createOverAllSheetLayout(btw_sheet)

    # COLOURING COLUMNS OF ABST AND COMP
    ws = wb["OverAll Result"]

    header_cols = ws.iter_cols(max_row=1, min_row=1, min_col=1)
    for head_col in header_cols:
        if str(head_col[0].value).lower() == "result":
            result_rows = ws.iter_rows(min_row=3, min_col=head_col[0].col_idx, max_col=head_col[0].col_idx)
            for result_ in result_rows:
                if str(result_[0].value) == "ABST":
                    to_color = ws.iter_cols(max_row=result_[0].row, min_row=result_[0].row, min_col=1)
                    for itm3 in to_color:
                        itm3[0].fill = PatternFill(start_color="ff7b63", end_color="ff7b63", fill_type="solid")
                elif str(result_[0].value) == "COMP":
                    to_color = ws.iter_cols(max_row=result_[0].row, min_row=result_[0].row, min_col=1)
                    for itm3 in to_color:
                        itm3[0].fill = PatternFill(start_color="ffe945", end_color="ffe945", fill_type="solid")
            break

    ws.merge_cells(start_row=len(df_.index)+3,end_row=len(df_.index)+3, start_column=1, end_column=3)
    ws.cell(row=len(df_.index)+3, column=1, value=f"Total Student : {len(df_.index)}").fill = PatternFill(fill_type="solid", start_color="ffc187", end_color="ffc187")
    ws.merge_cells(start_row=len(df_.index)+4,end_row=len(df_.index)+4, start_column=1, end_column=3)
    ws.cell(row=len(df_.index)+4, column=1, value=f"Subject Average(s)").fill = PatternFill(fill_type="solid", start_color="9bff87", end_color="9bff87")
    wb.save(export_file.name)
    messagebox.showinfo("Done", "File export successful!")

def subcode_to_subname(code_list):
    to_ret = []
    for itm in code_list:
        to_ret.append(subject_by_code[itm]["Name"])
    return to_ret

# -------------------- END -----------------------

header_canva = Canvas(root, bg="#FFFFFF", highlightthickness=2, highlightbackground="#000000")
def on_add_data_file():
    global selected_file, status_area_frame

    add_data_file_opt.config(text="Importing...", state="disabled")
    root.config(cursor='watch')
    root.update()
    selected_file = filedialog.askopenfile("r", filetypes=[("text files", "*.txt")])
    if selected_file:
        if status_area_frame:
            status_area_frame = status_area_frame.destroy()
        try:
            dataLineExtractor(selected_file)
        except BaseException as e:
            messagebox.showerror("Invalid File", f"File format is invalid please provide a valid file.\nerror : {e}")
    add_data_file_opt.config(text="Add Data File", state="normal")
    root.config(cursor='arrow')

add_data_file_opt = Button(header_canva, text="Add Data File", font=("Helvetica", 12), command=on_add_data_file)
add_data_file_opt.pack(padx=2, pady=2, side=LEFT)

def on_click_export():
    export_data_file_opt.config(text="Exporting...", state="disabled")
    root.config(cursor='watch')
    try:
        writeToExcel()
    except BaseException as e:
        messagebox.showerror("Failed", f"Error in exporting excel file...\nError : {e}")
    export_data_file_opt.config(text="Export Data File", state="normal")
    root.config(cursor='arrow')

export_data_file_opt = Button(header_canva, text="Export Data File", font=("Helvetica", 12), command=on_click_export)
export_data_file_opt.pack(padx=2, pady=2, side=LEFT)

def on_configs():
    configs_opt.config(state="disabled")

    config_canva = Canvas(content_left_canva, bg="#FFFFFF", highlightthickness=2, highlightbackground="#000000")
    def on_close_config():
        config_canva.destroy()
        configs_opt.config(state="normal")
    btn_close = Button(config_canva, text="Close", bg="#FF0000", fg="#FFFFFF", activebackground="#FF0000", activeforeground="#FFFFFF", command=on_close_config)
    btn_close.pack(anchor="e", pady=2)

    lbl_configs = Label(config_canva, text="Global Configurations", font=("Arial", 16, "bold", "underline"), bg="#FFFFFF")
    lbl_configs.pack(pady=2)

    # ---------------------- ADD SUBJECT EDITOR -------------------------
    sub_name_var = StringVar()
    sub_code_var = StringVar()

    frame_add_sub = Frame(config_canva, bg="#FFFFFF", highlightthickness=1, highlightbackground="#000000")
    lbl_add_sub = Label(frame_add_sub, text="Global Subject Code", bg="#FFFFFF", font=("Arial", 14))
    lbl_add_sub.pack(anchor="nw")

    frame_add_sub_E1 = Frame(frame_add_sub, bg="#FFFFFF")
    lbl_code = Label(frame_add_sub_E1, text="Subject Code", bg="#FFFFFF", font=("Arial", 12))
    lbl_code.pack(side=LEFT)
    entry_code = Entry(frame_add_sub_E1, highlightthickness=1, highlightcolor="#0000FF", textvariable=sub_code_var)
    entry_code.pack(side=LEFT)
    frame_add_sub_E1.pack()

    frame_add_sub_E2 = Frame(frame_add_sub, bg="#FFFFFF")
    lbl_name = Label(frame_add_sub_E2, text="Subject Name", bg="#FFFFFF", font=("Arial", 12))
    lbl_name.pack(side=LEFT)
    entry_name = Entry(frame_add_sub_E2, highlightthickness=1, highlightcolor="#0000FF", textvariable=sub_name_var)
    entry_name.pack(side=LEFT)
    frame_add_sub_E2.pack()

    frame_add_sub_o1 = Frame(frame_add_sub, bg="#FFFFFF")
    def on_add_sub():
        # subject_by_codes has already loaded the subject_code_refer file but that variable is editing some value in runtime so that below I again loading the file.
        all_subject_by_codes = json.load(open("subject_code_refer.json", "r"))
        if sub_code_var.get() in all_subject_by_codes:
            if not messagebox.askokcancel("Already exists", f"This code already exist with subject name {all_subject_by_codes[sub_code_var.get()]['Name']} do you like to change it with {sub_name_var.get()}?"):
                return
        all_subject_by_codes[sub_code_var.get()] = {"Name": sub_name_var.get()}
        json.dump(all_subject_by_codes, open("subject_code_refer.json", "w"))
        messagebox.showinfo("successful", "subject added successfully RESTART the software to commit changes.")

    add_sub = Button(frame_add_sub_o1, text="Add", bg="#90EE90", width=14, command=on_add_sub)
    add_sub.pack(side=LEFT)

    def on_remove_sub():
        # subject_by_codes has already loaded the subject_code_refer file but that variable is editing some value in runtime so that below I again loading the file.
        all_subject_by_codes = json.load(open("subject_code_refer.json", "r"))
        if sub_code_var.get() in all_subject_by_codes:
            del all_subject_by_codes[sub_code_var.get()]
            json.dump(all_subject_by_codes, open("subject_code_refer.json", "w"))
            messagebox.showinfo("successful", f"subject deleted successfully RESTART the software to commit changes.")
        else:
            messagebox.showerror("Not fount", f"No subject with code {sub_code_var.get()} found.")
    remove_sub = Button(frame_add_sub_o1, text="Remove", bg="#ffcccb", width=14, command=on_remove_sub)
    remove_sub.pack(side=LEFT)
    frame_add_sub_o1.pack()

    frame_add_sub.pack(padx=3, pady=2, anchor="w", ipadx=10)
    # -------------------------------- END ---------------------------------------


    # ------------------------- EDIT MAX MARKS CONSTANT -----------------------
    mm_text_var = IntVar(value=configs_["max_marks_1_subject"])
    frame_edit_mm = Frame(config_canva, bg="#FFFFFF", highlightthickness=1, highlightbackground="#000000")
    lbl_edit_mm = Label(frame_edit_mm, text="Global Max Marks Per Subject", bg="#FFFFFF", font=("Arial", 14))
    lbl_edit_mm.pack(anchor="nw")

    frame_edit_mm_E1 = Frame(frame_edit_mm, bg="#FFFFFF")
    lbl_max_marks = Label(frame_edit_mm_E1, text="Max Marks of Single Subject", bg="#FFFFFF", font=("Arial", 12))
    lbl_max_marks.pack(side=LEFT)
    entry_max_marks = Entry(frame_edit_mm_E1, highlightthickness=1, highlightcolor="#0000FF", textvariable=mm_text_var)
    entry_max_marks.pack(side=LEFT)
    frame_edit_mm_E1.pack()

    def on_save_mm():
        configs_["max_marks_1_subject"] = mm_text_var.get()
        json.dump(configs_, open("configs.json", "w"))
        messagebox.showinfo("Successful", f"Changed max marks to {configs_['max_marks_1_subject']}")

    save_mm =Button(frame_edit_mm, text="Save", bg="#90EE90", width=14, command=on_save_mm)
    save_mm.pack()

    frame_edit_mm.pack(padx=3, pady=2, anchor="w", ipadx=10)

    # --------------------------------- END ------------------------------


    # -------------------------- EDIT MAIN SUBJECT CODE ----------------
    main_sub_name = StringVar(value=configs_["main_subject_code"])

    frame_edit_MS = Frame(config_canva, bg="#FFFFFF", highlightthickness=1, highlightbackground="#000000")
    lbl_edit_MS = Label(frame_edit_MS, text="Global Main Subject", bg="#FFFFFF", font=("Arial", 14))
    lbl_edit_MS.pack(anchor="nw")

    frame_edit_MS_E1 = Frame(frame_edit_MS, bg="#FFFFFF")
    lbl_main_subject = Label(frame_edit_MS_E1, text="Main Subject", bg="#FFFFFF", font=("Arial", 12))
    lbl_main_subject.pack(side=LEFT)
    menu_subjects = OptionMenu(frame_edit_MS_E1, main_sub_name, *subject_by_code.keys())
    menu_subjects.pack(side=LEFT)
    frame_edit_MS_E1.pack()

    def on_save_mm():
        configs_["main_subject_code"] = main_sub_name.get()
        json.dump(configs_, open("configs.json", "w"))
        messagebox.showinfo("Successful", f"Changed main subject to {configs_['main_subject_code']}")

    save_mm = Button(frame_edit_MS, text="Save", bg="#90EE90", width=14, command=on_save_mm)
    save_mm.pack()

    frame_edit_MS.pack(padx=3, pady=2, anchor="w", ipadx=10)

    # ------------------------------- END ------------------------------

    # -------------------------- EDIT MAX BEST SUBJECT RANGE ----------------
    max_best_subs = IntVar(value=configs_["max_best_sub_range"])

    frame_edit_MBS = Frame(config_canva, bg="#FFFFFF", highlightthickness=1, highlightbackground="#000000")
    lbl_edit_MBS = Label(frame_edit_MBS, text="Global Max Best Subject Range", bg="#FFFFFF", font=("Arial", 14))
    lbl_edit_MBS.pack(anchor="nw")

    frame_edit_MBS_E1 = Frame(frame_edit_MBS, bg="#FFFFFF")
    lbl_best_subjects = Label(frame_edit_MBS_E1, text="Max Best Subject Range\n(include 1 main subject)", bg="#FFFFFF", font=("Arial", 12))
    lbl_best_subjects.pack(side=LEFT)
    entry_best_subjects = Entry(frame_edit_MBS_E1, highlightthickness=1, highlightcolor="#0000FF", textvariable=max_best_subs)
    entry_best_subjects.pack(side=LEFT)
    frame_edit_MBS_E1.pack()

    def on_save_mm():
        configs_["max_best_sub_range"] = max_best_subs.get()
        json.dump(configs_, open("configs.json", "w"))
        messagebox.showinfo("Successful", f"Changed max best subjects range to {configs_['max_best_sub_range']}")

    save_mm = Button(frame_edit_MBS, text="Save", bg="#90EE90", width=14, command=on_save_mm)
    save_mm.pack()

    frame_edit_MBS.pack(padx=3, pady=2, anchor="w", ipadx=10)
    # ------------------------------- END ------------------------------


    # ------------------------- Dominant subjects ----------------------
    dominant_subs_csv = StringVar(value=",".join(configs_["dominant_subjects"]))

    frame_edit_DS = Frame(config_canva, bg="#FFFFFF", highlightthickness=1, highlightbackground="#000000")
    lbl_edit_DS = Label(frame_edit_DS, text="Global Dominant Subject", bg="#FFFFFF", font=("Arial", 14))
    lbl_edit_DS.pack(anchor="nw")

    frame_edit_DS_E1 = Frame(frame_edit_DS, bg="#FFFFFF")
    lbl_dominant_subject = Label(frame_edit_DS_E1, text="Dominant Subjects\n(separated by ',')", bg="#FFFFFF", font=("Arial", 12))
    lbl_dominant_subject.pack(side=LEFT)
    dominant_subject = Entry(frame_edit_DS_E1, textvariable=dominant_subs_csv, highlightthickness=1, highlightcolor="#0000FF")
    dominant_subject.pack(side=LEFT)
    frame_edit_DS_E1.pack()

    def on_save_mm():
        configs_["dominant_subjects"] = list_formatter(dominant_subs_csv.get().replace(" ", "").split(","))
        for cds in configs_["dominant_subjects"]:
            if cds not in subject_by_code:
                messagebox.showerror("Failed", f"Subject code '{cds}' not exists please add this subject code first.")
                return
        json.dump(configs_, open("configs.json", "w"))
        messagebox.showinfo("Successful", f"Successfully edited dominant subjects.")

    save_DS = Button(frame_edit_DS, text="Save", bg="#90EE90", width=14, command=on_save_mm)
    save_DS.pack()

    frame_edit_DS.pack(padx=3, pady=2, anchor="w", ipadx=10)
    # ------------------------- END ------------------------------------

    config_canva.place(x=0, y=0, relwidth=1.0, relheight=1.0)

configs_opt = Button(header_canva, text="Config", font=("Helvetica", 12), command=on_configs)
configs_opt.pack(padx=2, pady=2, side=LEFT)


def on_click_logs():
    global logs_string
    logs_opt.config(state="disabled")
    logs_opt.config(background="#d9d9d9", foreground="#000000")
    log_canvas = Canvas(content_left_canva, bg="#FFFFFF", highlightthickness=2, highlightbackground="#000000")

    fr_close_btn = Frame(log_canvas, bg="#FFFFFF")
    lbl_logs = Label(fr_close_btn, text="LOGS", font=("Helvetica", 16, "bold", "italic"), background="#FFFFFF")
    lbl_logs.pack(side=LEFT, padx=5)
    def onClickCloseLogs():
        log_canvas.destroy()
        logs_opt.config(state="normal")

    close_btn = Button(fr_close_btn, text="Close", foreground="#FFFFFF", background="#db2518", command=onClickCloseLogs)
    close_btn.pack(side=RIGHT, padx=1)
    def onClearLogs():
        global logs_string
        logs_string = ""
        onClickCloseLogs()
        on_click_logs()
        messagebox.showinfo("Cleared", "Logs have been cleared.")
    clearLog_btn = Button(fr_close_btn, text="Clear", foreground="#FFFFFF", background="#0000FF", command=onClearLogs)
    clearLog_btn.pack(side=RIGHT, padx=1)
    fr_close_btn.pack(fill=X, padx=3, pady=3)

    text_ar = Text(log_canvas, highlightthickness=1, highlightcolor="#000000", bg="#FFFFFF", fg="#000000", font=("Arial", 12))
    text_ar.insert(END, logs_string)
    text_ar.config(state="disabled")
    text_ar.focus()
    text_ar.pack(fill=BOTH, padx=3, expand=True, pady=3)

    log_canvas.place(x=0, y=0, relwidth=1.0, relheight=1.0)

logs_opt = Button(header_canva, text="Logs", font=("Helvetica", 12), command=on_click_logs)
logs_opt.pack(padx=2, pady=2, side=LEFT)

def on_click_about():
    about_opt.config(state="disabled")

    text_to_show = """
    ------------------------------- Function of software --------------------------------
    This software is personally build for sort and structuring data of CBSE board examinations result of students. It requires a txt file 
    which contain data in a format describe below then this software will sort students, their gender, their roll number and
    percentage average etc. Sorted data can be exported as an Excel or JSON file.
    
    ------------------------------- Data format required --------------------------------
    DATA_LINE1 :- ROLL_NO GENDER STUDENT_NAME SUB_CODE1 SUB_CODE2 SUB_CODE3......SUB_CODE(n) GRADE1 GRADE2 GRADE3 RESULT COMPARTMENT_SUBJECT_CODES
    DATA_LINE2 :- SUB1_MARKS SUB1_GRADE SUB2_MARKS SUB2_GRADE ...... SUB(n)_MARKS SUB(n)_GRADE
    
    > how should be DATA_LINE1 ?
    1. ROLL_NO should be in range of 6 to 10 digits.
    2. Gender M for male, F for female
    3. NAME can contain white spaces.
    4. SUB_CODE1 to infinity works just remember that DATA_LINE2 should contain marks and grade of SUB_CODE(n) in proper format.
    5. (optional) GRADE1 GRADE2 GRADE3 not actually required software works same if its present of not.
    6. RESULT it should be FAIL or PASS or COMP or ABST and its required.

    > how should be DATA_LINE2
    1. SUB(n)_MARKS and SUB(n)_GRADE represents the marks and grade with respect to SUB_CODE(n).
    
    Note : DATA_LINE1 and DATA_LINE2 both are required if any of one missing so software will delete that entry and proceed for next proper paired data lines.
    
    ------------------------------- Configurations --------------------------------
    1. Global Subject Code : Here you can add new subject code and its name in software.
        
    2. Global Max Marks Per Subject : It represents the max marks of which paper held. Like a student got 40 out of 80 so 80
    is the max marks. It will be same for all subjects.
    
    3. Global Main Subject : It represents the main subject which is common in all the sides for example english is language
    subject which is same for all the fields (science, arts etc.).
    
    4. Global Max Best Subject Range : It represents how many top best subject marks of a student will be consider for
    calculate variable columns like average(n), percentage(n) etc. its value should be no_of_best_subjects + 1
    for example : 4 + 1 = 5 where no_of_best_subjects = 4 and the 1 here for english/main subject so value of GLOBAL will be 5.
    
    5. Global Dominant Subjects : Its helps to control the column format of table like if we want subject format english
    then Physics chemistry .... so we can use this entry to do that. Just provide subject codes and seperate them
    using ','.
    """

    about_canva = Canvas(content_left_canva, bg="#FFFFFF", highlightthickness=2, highlightbackground="#000000")
    def on_close_about():
        about_canva.destroy()
        about_opt.config(state="normal")
    btn_close = Button(about_canva, text="Close", bg="#FF0000", fg="#FFFFFF", activebackground="#FF0000", activeforeground="#FFFFFF", command=on_close_about)
    btn_close.pack(anchor="e", pady=2)

    lbl_about_head = Label(about_canva, text="About", bg="#FFFFFF", font=("Arial", 18, "bold", "underline"))
    lbl_about_head.pack()

    text_ar = Text(about_canva, highlightthickness=1, highlightcolor="#000000", bg="#FFFFFF", fg="#000000", font=("Arial", 12))
    text_ar.insert(END, text_to_show)
    text_ar.config(state="disabled")
    text_ar.focus()
    text_ar.pack(fill=BOTH, padx=3, expand=True)

    def on_developer():
        webbrowser.open("https://github.com/PowerPizza")

    btn_developer = Button(about_canva, text="DEVELOPER", bg="blue", fg="yellow", command=on_developer)
    btn_developer.pack(anchor="e", pady=3, padx=3)

    about_canva.place(x=0, y=0, relwidth=1.0, relheight=1.0)

about_opt = Button(header_canva, text="About", font=("Helvetica", 12), command=on_click_about)
about_opt.pack(padx=2, pady=2, side=LEFT)

label_imported_file = Label(header_canva, text="", font=("Helvetica", 12), bg="#FFFFFF")
label_imported_file.pack(side=LEFT, pady=2, padx=2)
header_canva.pack(fill=X, ipadx=2, ipady=2, padx=2, pady=1)

content_left_canva = Canvas(root, bg="#FFFFFF", highlightthickness=2, highlightbackground="#000000")

def begin_status(selected_file_name):
    global status_area_frame

    search_by_val = StringVar(value="Roll No.")
    search_for_val = StringVar()
    page_indexer = 0

    frame_status_area = Frame(content_left_canva, bg="#FFFFFF")
    status_area_frame = frame_status_area

    lbl_file_name = Label(frame_status_area, text=f"File {selected_file_name}", bg="#FFFFFF", font=("Arial", 14))
    lbl_file_name.pack()

    lbl_total_data = Label(frame_status_area, text=f"Total Data : {len(df_.index)}", bg="#FFFFFF", font=("Arial", 14))
    lbl_total_data.pack()

    frame_search_area = Frame(frame_status_area, bg="#fcffdb", highlightthickness=1, highlightbackground="#000000")
    search_entry = Entry(frame_search_area, font=('Helvetica', 14), border=2, relief='ridge', textvariable=search_for_val)
    search_entry.pack(side=LEFT, fill=BOTH, expand=True)
    search_opt_menu = OptionMenu(frame_search_area, search_by_val, *["Roll No.", "Student Name"])
    search_opt_menu.config(font=('Helvetica', 14))
    search_opt_menu.pack(side=LEFT)
    def on_search_():
        if search_by_val.get() == "Roll No.":
            if len(search_for_val.get()):
                idx = df_.loc[df_["roll_no"].astype(str) == search_for_val.get()].index
                if len(idx):
                    display_entry_values(idx[0])
                else:
                    messagebox.showinfo("Not Found", "No search results found!")

        elif search_by_val.get() == "Student Name":
            search_btn.config(state="disabled")
            search_results = df_.loc[df_["student_name"].str.startswith(search_for_val.get())].index
            multiple_result_win = Toplevel(root)
            multiple_result_win.transient(root)
            multiple_result_win.title(f"Search Results")
            def close_search_result():
                search_btn.config(state="normal")
                multiple_result_win.destroy()

            multiple_result_win.protocol("WM_DELETE_WINDOW", close_search_result)

            lbl_result_found = Label(multiple_result_win, text=f"`{len(search_results)}` Search Result Found", font=("Helvetica", 16, "bold"))
            lbl_result_found.pack(anchor=W)

            scroll_bar = Scrollbar(multiple_result_win)
            scroll_bar.pack(side=RIGHT, fill=Y)
            search_list = Listbox(multiple_result_win, yscrollcommand=scroll_bar.set)
            idx = 0
            for result in search_results:
                print(result)
                search_list.insert(idx, f"{df_.iloc[result]['roll_no']} -- {df_.iloc[result]['student_name']}")
                idx += 1
            search_list.pack(anchor=W, fill=BOTH, expand=True, padx=2, pady=2)
            def on_select_result(*eve):
                display_entry_values(search_results[search_list.curselection()[0]])

            search_list.bind("<<ListboxSelect>>", on_select_result)
            scroll_bar.config(command=search_list.yview)

            multiple_result_win.mainloop()

    search_btn = Button(frame_search_area, text="Search", font=('Helvetica', 14), command=on_search_)
    search_btn.pack(side=RIGHT)
    frame_search_area.pack(fill=X, anchor=NE, padx=12, pady=2)

    result_area = Frame(frame_status_area, bg="#FFFFFF", highlightthickness=1, highlightbackground="#000000")
    def entry_creator( txt_, value_, master=result_area):
        text_var = StringVar(value=value_)
        frame_e1 = Frame(master, bg="#FFFFFF")
        lbl_entry = Label(frame_e1, text=txt_, font=("Helvetica", 12), bg="#FFFFFF")
        lbl_entry.pack(side=LEFT, padx=2)
        entry_e1 = Entry(frame_e1, textvariable=text_var, font=("Helvetica", 12), highlightthickness=1, highlightbackground="#000000")
        entry_e1.pack(side=LEFT, padx=2)
        frame_e1.pack(pady=2, fill=X)
        return {"text_var": text_var, "frame_": frame_e1}

    lbl_page = Label(result_area, text="Data Page : 0", font=("Helvetica", 12, "bold"), bg="#FFFFFF")
    lbl_page.pack(anchor="ne")

    set_roll_no = entry_creator("Roll. No.", "")["text_var"]
    set_name = entry_creator("Name", "")["text_var"]
    set_gender = entry_creator("Gender", "")["text_var"]

    frame_mg_pri = Frame(result_area, bg="#FFFFFF")
    frame_mg_scd = Frame(frame_mg_pri, bg="#FFFFFF")
    frame_mg_pri.pack(fill=X)

    set_total_marks = entry_creator("Marks Total", "")["text_var"]
    set_percentage_marks = entry_creator("Percentage", "")["text_var"]
    set_average_marks = entry_creator("Average", "")["text_var"]
    set_lowest_marks = entry_creator("Lowest", "")["text_var"]
    set_highest_marks = entry_creator("Highest", "")["text_var"]

    def display_entry_values(index_):
        nonlocal frame_mg_scd
        frame_mg_scd.destroy()
        frame_mg_scd = Frame(frame_mg_pri, bg="#FFFFFF")
        frame_mg_scd.pack(fill=X)

        marks_refr = df_.iloc[index_]["MARKS"]

        for c1 in range(0, len(unique_sub_codes), 2):
            fr2 = Frame(frame_mg_scd)
            if len(marks_refr.iloc[c1]) > 0:
                entry_creator(master=fr2, txt_=subject_by_code[unique_sub_codes[c1]]["Name"], value_=marks_refr.iloc[c1])["frame_"].pack(side=LEFT)
            if c1+1 < len(unique_sub_codes):
                if len(marks_refr.iloc[c1+1]) > 0:
                    entry_creator(master=fr2, txt_=subject_by_code[unique_sub_codes[c1+1]]["Name"], value_=marks_refr.iloc[c1+1])["frame_"].pack(side=LEFT)
            fr2.pack(fill=X, expand=True)

        set_roll_no.set(df_.iloc[index_]['roll_no'])
        set_name.set(df_.iloc[index_]['student_name'])
        set_gender.set(df_.iloc[index_]['gender'])
        set_total_marks.set(df_.iloc[index_]['Total OA'])
        set_percentage_marks.set(df_.iloc[index_]['Percentage OA'])
        set_average_marks.set(df_.iloc[index_]['Average OA'])

        highest_ = {"sub": None, "marks": 0}
        lowest_ = {"sub": None, "marks": 0}
        for c2 in range(len(unique_sub_codes)):
            if c2 == 0:
                lowest_["marks"] = int(marks_refr.iloc[c2])
                lowest_["sub"] = subject_by_code[unique_sub_codes[c2]]["Name"]

            if str(marks_refr.iloc[c2]).isnumeric() and int(marks_refr.iloc[c2]) > highest_["marks"]:
                highest_["marks"] = int(marks_refr.iloc[c2])
                highest_["sub"] = subject_by_code[unique_sub_codes[c2]]["Name"]

            elif str(marks_refr.iloc[c2]).isnumeric() and int(marks_refr.iloc[c2]) < lowest_["marks"]:
                lowest_["marks"] = int(marks_refr.iloc[c2])
                lowest_["sub"] = subject_by_code[unique_sub_codes[c2]]["Name"]
        set_highest_marks.set(highest_["sub"])
        set_lowest_marks.set(lowest_["sub"])

        lbl_page.config(text=f"Data Page : {index_+1}")
    def on_prev():
        nonlocal page_indexer
        button_next.config(state="normal")
        if page_indexer-1 < 0:
            button_prev.config(state="disabled")
            return
        page_indexer -= 1
        display_entry_values(page_indexer)

    button_prev = Button(result_area, text="Previous", font=("Helvetica", 16), command=on_prev)
    button_prev.pack(side=LEFT, padx=1, pady=1)

    def on_next():
        nonlocal page_indexer
        button_prev.config(state="normal")
        page_indexer += 1
        display_entry_values(page_indexer)
        if page_indexer + 1 == len(df_.index):
            button_next.config(state="disabled")

    button_next = Button(result_area, text="Next", font=("Helvetica", 16), command=on_next)
    button_next.pack(side=RIGHT, padx=1, pady=1)

    display_entry_values(0)

    result_area.pack(side=LEFT, anchor="nw", padx=12, fill=X, expand=True)

    frame_status_area.pack(fill=BOTH, side=LEFT, expand=True, padx=2, pady=2)

    return frame_status_area

content_left_canva.pack(side=LEFT, fill=BOTH, expand=True, padx=2, pady=1)

root.mainloop()